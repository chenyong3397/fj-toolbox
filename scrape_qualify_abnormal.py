#!/usr/bin/env python3
"""
福建省建设行业平台 - 质量检测企业资质异常自动统计程序
功能：
  1. 自动抓取质量检测企业资质异常数据
  2. 支持定时运行（通过 --schedule 参数或 Windows 任务计划）
  3. 支持手动运行（双击 run_now.bat 或在命令行运行）
  4. 自动保存带时间戳的结果
  5. 检测变化并生成对比报告

用法：
  python scrape_qualify_abnormal.py                    # 立即运行一次
  python scrape_qualify_abnormal.py --schedule         # 使用内置调度器
  python scrape_qualify_abnormal.py --config config.ini  # 指定配置文件
  python setup_config.py                               # 交互式配置向导
"""

import requests, urllib3, json, io, base64, time, random, os, sys, argparse, logging, configparser, re
from PIL import Image
import numpy as np

# 强制 UTF-8 输出，避免 Windows CMD/GBK 编码错误
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from urllib.parse import quote
from datetime import datetime, timedelta

# 导入通知模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import notify

urllib3.disable_warnings()

# ============================================================
# 配置区（可通过 config.json 覆盖）
# ============================================================
DEFAULT_CONFIG = {
    "platform": {
        "base_url": "https://220.160.52.164:8813/credit",
        "index_url": "https://220.160.52.164:8813/gaia/infoPublic/index.html"
    },
    "query": {
        "industry_id": 2,
        "page_size": 30,
        "request_delay_min": 0.3,
        "request_delay_max": 0.8,
        "refresh_token_every": 25
    },
    "output": {
        "output_dir": "output",
        "filename_prefix": "质量检测_资质异常",
        "keep_history": 30
    },
    "schedule": {
        "enabled": False,
        "time": "08:00",
        "weekdays_only": False
    },
    "notification": {
        "enabled": False,
        "on_change_only": False,
        "serverchan": {
            "enabled": False,
            "sendkey": ""
        },
        "wxpusher": {
            "enabled": False,
            "app_token": "",
            "topic_ids": [],
            "uids": []
        }
    }
}

CREDIT = DEFAULT_CONFIG["platform"]["base_url"]
INDEX = DEFAULT_CONFIG["platform"]["index_url"]

# ============================================================
# 工具函数
# ============================================================

def load_config(config_path="config.ini"):
    """加载配置文件（支持 INI 和 JSON），与默认配置合并"""
    config = DEFAULT_CONFIG.copy()
    
    if not os.path.exists(config_path):
        # 尝试 .ini 和 .json
        base = os.path.splitext(config_path)[0]
        for ext in ['.ini', '.json']:
            alt_path = base + ext
            if os.path.exists(alt_path):
                config_path = alt_path
                break
    
    if os.path.exists(config_path):
        try:
            if config_path.endswith('.json'):
                # JSON 格式（兼容旧版）
                with open(config_path, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                for key in user_config:
                    if key in config and isinstance(config[key], dict):
                        config[key].update(user_config[key])
                    else:
                        config[key] = user_config[key]
            else:
                # INI 格式（默认）
                cp = configparser.ConfigParser()
                cp.read(config_path, encoding='utf-8')
                
                def _bool(v):
                    return v.strip().lower() == 'true'
                
                # 需要合并到 notification 的子 section
                def _read_section(section, dest, key_types):
                    if cp.has_section(section):
                        for key, typ in key_types.items():
                            if cp.has_option(section, key):
                                val = cp.get(section, key)
                                try:
                                    config[dest][key] = typ(val)
                                except:
                                    pass
                
                # 顶层 section（直接对应 config 中的 key）
                _read_section('platform', 'platform',
                    {'base_url': str, 'index_url': str})
                _read_section('query', 'query',
                    {'industry_id': int, 'page_size': int,
                     'request_delay_min': float, 'request_delay_max': float,
                     'refresh_token_every': int})
                _read_section('output', 'output',
                    {'output_dir': str, 'filename_prefix': str, 'keep_history': int})
                _read_section('schedule', 'schedule',
                    {'enabled': _bool, 'time': str, 'weekdays_only': _bool})
                _read_section('notification', 'notification',
                    {'enabled': _bool, 'on_change_only': _bool})
                
                # serverchan 和 wxpusher 合并到 notification 下
                _read_section('serverchan', 'notification',
                    {'enabled': _bool})
                _read_section('wxpusher', 'notification',
                    {'enabled': _bool})
                if cp.has_section('serverchan') and cp.has_option('serverchan', 'sendkey'):
                    config['notification']['serverchan'] = {
                        'enabled': _bool(cp.get('serverchan', 'enabled')),
                        'sendkey': cp.get('serverchan', 'sendkey')
                    }
                if cp.has_section('wxpusher') and cp.has_option('wxpusher', 'app_token'):
                    def _parse_list(val):
                        """解析逗号分隔的列表"""
                        return [x.strip() for x in val.split(',') if x.strip()]
                    wx_config = {
                        'enabled': _bool(cp.get('wxpusher', 'enabled')),
                        'app_token': cp.get('wxpusher', 'app_token'),
                    }
                    if cp.has_option('wxpusher', 'topic_ids') and cp.get('wxpusher', 'topic_ids').strip():
                        wx_config['topic_ids'] = _parse_list(cp.get('wxpusher', 'topic_ids'))
                    else:
                        wx_config['topic_ids'] = []
                    if cp.has_option('wxpusher', 'uids') and cp.get('wxpusher', 'uids').strip():
                        wx_config['uids'] = _parse_list(cp.get('wxpusher', 'uids'))
                    else:
                        wx_config['uids'] = []
                    config['notification']['wxpusher'] = wx_config
                    
            print(f"  [OK] 已加载配置文件: {config_path}")
        except Exception as e:
            print(f"  [WARN] 配置文件加载失败，使用默认配置: {e}")
    else:
        # 创建默认 INI 配置文件
        _create_default_ini(config_path)
        print(f"  [OK] 已创建默认配置文件: {config_path}")
        print(f"    用记事本打开编辑: notepad {os.path.abspath(config_path)}")
    
    return config


def _create_default_ini(path):
    """创建默认 INI 配置文件"""
    with open(path, 'w', encoding='utf-8') as f:
        f.write("""; ============================================================
; 质量检测企业资质异常自动统计程序 - 配置文件
; 用记事本打开编辑即可，保存为 UTF-8 编码
; 也可以运行 setup_config.py 交互式配置
; ============================================================

[platform]
base_url = https://220.160.52.164:8813/credit
index_url = https://220.160.52.164:8813/gaia/infoPublic/index.html

[query]
industry_id = 2
page_size = 30
request_delay_min = 0.3
request_delay_max = 0.8
refresh_token_every = 25

[output]
output_dir = output
filename_prefix = 质量检测_资质异常
keep_history = 30

[schedule]
enabled = false
time = 08:00
weekdays_only = false

[notification]
enabled = true
on_change_only = false

[serverchan]
enabled = true
sendkey = 在此填入你的Server酱SendKey

[wxpusher]
enabled = true
app_token = 在此填入你的WxPusher appToken
topic_ids = 在此填入主题ID（数字），多个用英文逗号分隔
uids = 在此填入用户UID（字符串），多个用英文逗号分隔
""")



def setup_logging(output_dir):
    """设置日志"""
    os.makedirs(output_dir, exist_ok=True)
    log_file = os.path.join(output_dir, f"run_{datetime.now().strftime('%Y%m%d')}.log")
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)


def make_session():
    """创建会话（通过 Cloudflare Worker 代理）"""
    from proxy_utils import ProxySession
    s = ProxySession()
    s.verify = False
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Referer': INDEX,
        'Origin': 'https://220.160.52.164:8813',
    })
    s.get(INDEX, timeout=15)
    return s


def solve_captcha(s):
    """破解滑块验证码，返回 token"""
    for attempt in range(8):
        try:
            r = s.get(f'{CREDIT}/captcha', timeout=15)
            j = r.json()
            if j.get('code') != 1:
                time.sleep(1)
                continue
            cap = j['data']
            tok = cap['token']
            ct = cap.get('captchaType', 0)
            
            # 使用 alpha 通道模板匹配
            bg = Image.open(io.BytesIO(base64.b64decode(cap['imageBase64']))).convert('L')
            sl = Image.open(io.BytesIO(base64.b64decode(cap['sliderBase64']))).convert('RGBA')
            
            bg_f = np.array(bg, dtype=np.float32)
            sl_f = np.array(sl, dtype=np.float32)
            alpha = sl_f[:, :, 3] / 255.0
            
            sl_y = cap.get('sliderY', 0)
            sl_h = alpha.shape[0]
            sl_w = alpha.shape[1]
            img_w = bg_f.shape[1]
            
            roi = bg_f[max(0, sl_y):min(bg_f.shape[0], sl_y + sl_h), :]
            scores = []
            for x in range(img_w - sl_w):
                ra = alpha[:roi.shape[0], :min(sl_w, img_w - x)]
                rr = roi[:ra.shape[0], x:x + ra.shape[1]]
                scores.append((rr * ra).sum() / max(ra.sum(), 1e-6))
            best_x = int(np.argmin(scores))
            
            s.headers['Content-Type'] = 'application/x-www-form-urlencoded'
            s.headers['captcha-token'] = tok
            r2 = s.post(f'{CREDIT}/captcha?captchaType={ct}&code={best_x}', timeout=15)
            s.headers.pop('Content-Type', None)
            s.headers.pop('captcha-token', None)
            
            r2j = r2.json()
            if r2j.get('code') == 1 or r2j.get('data') == True:
                return tok
        except Exception as e:
            print(f'  [验证码错误] {e}')
            time.sleep(1)
    return None


def api_get(s, token, url, params=None, max_retry=5):
    """带自动验证码刷新的 GET 请求"""
    current_token = token
    for attempt in range(max_retry):
        try:
            headers = {'captcha-token': current_token} if current_token else {}
            r = s.get(url, params=params, headers=headers, timeout=30)
            result = r.json()
            code = result.get('code')
            if code == 1:
                return result.get('data'), None, current_token
            elif code in (1001, 3001):
                print(f'    [令牌过期] 重新破解验证码... (尝试 {attempt + 1})')
                new_tok = solve_captcha(s)
                if new_tok:
                    current_token = new_tok
                    continue
                return None, 'captcha_failed', current_token
            else:
                msg = result.get('message', result.get('msg', ''))
                return None, f'code={code} msg={msg}', current_token
        except Exception as e:
            if attempt < max_retry - 1:
                print(f'    [请求错误] {e}, 重试中... (尝试 {attempt + 1})')
                time.sleep(2)
                new_tok = solve_captcha(s)
                if new_tok:
                    current_token = new_tok
            else:
                return None, str(e), current_token
    return None, 'max_retry', current_token


# ============================================================
# 核心抓取逻辑
# ============================================================

def run_scrape(config):
    """执行完整的抓取流程，返回结果字典"""
    print('=' * 70)
    print('  福建省建设行业平台 - 质量检测企业资质异常统计')
    print('=' * 70)
    
    s = make_session()
    CREDIT = config["platform"]["base_url"]
    
    # ---- 阶段1：获取质量检测企业名单 ----
    print('\n[阶段 1/4] 获取质量检测企业名单...')
    token = solve_captcha(s)
    if not token:
        print('  [FAIL] 致命错误：无法破解验证码')
        return None
    print('  [OK] 验证码破解成功')
    
    all_companies = []
    page = 1
    page_size = config["query"]["page_size"]
    while True:
        params = {
            'industryId': str(config["query"]["industry_id"]),
            'outside': 'false',
            'pageNum': str(page),
            'pageSize': str(page_size)
        }
        data, err, token = api_get(s, token, f'{CREDIT}/tendering/corps', params)
        if err:
            print(f'  第 {page} 页错误: {err}')
            break
        records = data.get('list', [])
        all_companies.extend(records)
        total = data.get('total', 0)
        print(f'  第 {page} 页: {len(records)} 家企业，累计 {len(all_companies)}/{total}')
        if page * page_size >= total:
            break
        page += 1
        time.sleep(random.uniform(
            config["query"]["request_delay_min"],
            config["query"]["request_delay_max"]
        ))
    
    print(f'\n  共获取 {len(all_companies)} 家质量检测企业')

    # 过滤掉名称含"测试"的企业（测试用账号）
    all_companies = [c for c in all_companies
                     if '测试' not in (c.get('corpName', '') or c.get('companyName', ''))]
    print(f'  过滤后剩余 {len(all_companies)} 家（已排除名称含"测试"的企业）')

    # ---- 阶段2：逐家检查资质异常 ----
    print('\n[阶段 2/4] 逐家检查企业资质异常...')
    abnormal_companies = []
    normal_count = 0
    error_count = 0
    refresh_every = config["query"]["refresh_token_every"]
    
    for i, comp in enumerate(all_companies):
        corp_name = comp.get('corpName', '') or comp.get('companyName', '')
        sc = comp.get('socialCreditCode', '')
        
        if not corp_name:
            print(f'  [{i + 1}/{len(all_companies)}] 跳过：企业名称为空')
            continue
        
        # 过滤掉名称含"测试"的企业（测试用账号）
        if '测试' in corp_name:
            print(f'  [{i + 1}/{len(all_companies)}] 跳过：测试企业 {corp_name}')
            continue
        
        encoded_name = quote(corp_name, safe='')
        url = f'{CREDIT}/tendering/qualify/{encoded_name}'
        data, err, token = api_get(s, token, url, {'pageNum': '1', 'pageSize': '30'})
        
        if err:
            print(f'  [{i + 1}/{len(all_companies)}] 错误: {corp_name} -> {err}')
            error_count += 1
            if 'captcha' in str(err) or '1001' in str(err):
                token = solve_captcha(s)
            time.sleep(1)
            continue
        
        # 解析资质列表
        qualify_records = []
        if isinstance(data, dict):
            qualify_records = data.get('list', [])
        elif isinstance(data, list):
            qualify_records = data
        
        has_abnormal = False
        abnormal_items = []
        
        for qrec in qualify_records:
            qualify_list = qrec.get('qualifyList', [])
            for qitem in qualify_list:
                if qitem.get('abnormal') == True:
                    has_abnormal = True
                    abnormal_items.append({
                        'qualifyNumber': qrec.get('qualifyNumber', ''),
                        'qualifyName': qitem.get('name', ''),
                        'qualifyAssociationId': qitem.get('qualifyAssociationId', ''),
                        'stateName': qrec.get('stateName', ''),
                        'officerName': qrec.get('officerName', ''),
                        'issuingTime': qrec.get('issuingTime', ''),
                        'startTime': qrec.get('startTime', ''),
                        'periodTime': qrec.get('periodTime', ''),
                    })
        
        if has_abnormal:
            abnormal_companies.append({
                'companyName': corp_name,
                'socialCreditCode': sc,
                'abnormalItems': abnormal_items,
            })
            print(f'  [{i + 1}/{len(all_companies)}] [WARN] 异常: {corp_name} ({len(abnormal_items)} 项)')
        else:
            normal_count += 1
            if (i + 1) % 20 == 0:
                print(f'  [{i + 1}/{len(all_companies)}] 进度: {normal_count} 正常, {len(abnormal_companies)} 异常, {error_count} 错误')
        
        # 定期刷新令牌
        if (i + 1) % refresh_every == 0:
            token = solve_captcha(s)
        
        time.sleep(random.uniform(
            config["query"]["request_delay_min"],
            config["query"]["request_delay_max"]
        ))
    
    print(f'\n[阶段 2/4] 完成:')
    print(f'  正常: {normal_count}')
    print(f'  异常: {len(abnormal_companies)}')
    print(f'  错误: {error_count}')
    
    # ---- 阶段3：获取异常资质详情 ----
    if abnormal_companies:
        print('\n[阶段 3/4] 获取异常资质详情...')
        token = solve_captcha(s)
        
        for i, comp in enumerate(abnormal_companies):
            sc = comp['socialCreditCode']
            for j, item in enumerate(comp['abnormalItems']):
                qaid = item.get('qualifyAssociationId', '')
                if not qaid:
                    continue
                
                url = f'{CREDIT}/tendering-extract/qualify'
                params = {'socialCreditCode': sc, 'qualifyAssociationId': str(qaid)}
                data, err, token = api_get(s, token, url, params)
                
                if err:
                    print(f'  [{i + 1}/{len(abnormal_companies)}] 详情错误: {comp["companyName"]} -> {err}')
                    item['detail'] = None
                else:
                    item['detail'] = data
                
                time.sleep(random.uniform(0.3, 0.5))
            
            if (i + 1) % 10 == 0:
                token = solve_captcha(s)
    
    # ---- 组装结果 ----
    result = {
        'scrape_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'total_companies': len(all_companies),
        'normal_count': normal_count,
        'abnormal_count': len(abnormal_companies),
        'error_count': error_count,
        'abnormal_companies': abnormal_companies,
    }
    
    print(f'\n[阶段 3/4] 完成')
    return result


# ============================================================
# Excel 生成
# ============================================================

def generate_excel(result, config):
    """生成 Excel 报告"""
    print('\n[阶段 4/4] 生成 Excel 报告...')
    
    output_dir = config["output"]["output_dir"]
    prefix = config["output"]["filename_prefix"]
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f'{prefix}_{timestamp}.xlsx')
    
    wb = openpyxl.Workbook()
    
    # 样式
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    abnormal_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    abnormal_companies = result['abnormal_companies']
    
    # Sheet 1: 资质异常明细
    ws1 = wb.active
    ws1.title = '资质异常明细'
    headers1 = ['序号', '企业名称', '统一社会信用代码', '资质证书编号', '异常资质名称',
                '核查批次', '核查单位', '核查开始时间', '核查截止时间', '发证时间', '有效期至']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 写入 Sheet 1 数据（记录每个企业的行范围用于合并）
    row_idx = 2
    company_ranges = []  # [(start_row, end_row), ...]
    sn = 0
    for comp in abnormal_companies:
        start_row = row_idx
        for item in comp['abnormalItems']:
            detail = item.get('detail') or {}
            sn += 1
            ws1.cell(row=row_idx, column=1, value=sn).border = thin_border
            ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=2, value=comp['companyName']).border = thin_border
            ws1.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=3, value=comp.get('socialCreditCode', '')).border = thin_border
            ws1.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=4, value=item.get('qualifyNumber', '')).border = thin_border
            ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=5, value=item.get('qualifyName', '')).border = thin_border
            ws1.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=6, value=detail.get('name', '')).border = thin_border
            ws1.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=7, value=detail.get('checkOfficerName', '')).border = thin_border
            ws1.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=8, value=detail.get('releaseTime', '')).border = thin_border
            ws1.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=9, value=detail.get('checkEndTime', '')).border = thin_border
            ws1.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=10, value=item.get('issuingTime', '')).border = thin_border
            ws1.cell(row=row_idx, column=10).alignment = Alignment(horizontal='center', vertical='center')
            ws1.cell(row=row_idx, column=11, value=item.get('periodTime', '')).border = thin_border
            ws1.cell(row=row_idx, column=11).alignment = Alignment(horizontal='center', vertical='center')
            for col in range(1, 12):
                ws1.cell(row=row_idx, column=col).fill = abnormal_fill
            row_idx += 1
        company_ranges.append((start_row, row_idx - 1))

    # 合并同一企业名称的 "序号"、"企业名称"、"统一社会信用代码" 列
    for start, end in company_ranges:
        if end > start:
            ws1.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws1.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
            ws1.merge_cells(start_row=start, start_column=3, end_row=end, end_column=3)
    
    # Sheet 2: 按企业汇总
    ws2 = wb.create_sheet('按企业汇总')
    headers2 = ['序号', '企业名称', '统一社会信用代码', '异常资质项数', '异常资质列表']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    for i, comp in enumerate(abnormal_companies):
        row = i + 2
        items = comp['abnormalItems']
        qualify_names = '；'.join(set(item.get('qualifyName', '') for item in items))
        ws2.cell(row=row, column=1, value=i + 1).border = thin_border
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=row, column=2, value=comp['companyName']).border = thin_border
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=row, column=3, value=comp['socialCreditCode']).border = thin_border
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=row, column=4, value=len(items)).border = thin_border
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws2.cell(row=row, column=5, value=qualify_names).border = thin_border
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(1, 6):
            ws2.cell(row=row, column=col).fill = abnormal_fill
    
    # Sheet 3: 统计概览
    ws3 = wb.create_sheet('统计概览')
    stats_data = [
        ['指标', '数值'],
        ['数据来源', '福建省建设行业信息公开平台'],
        ['查询范围', '省内企业 → 质量检测 → 资质异常'],
        ['质量检测企业总数', result['total_companies']],
        ['资质异常企业数', result['abnormal_count']],
        ['资质异常项数', sum(len(c['abnormalItems']) for c in abnormal_companies)],
        ['正常企业数', result['normal_count']],
        ['查询失败企业数', result['error_count']],
        ['异常企业占比', f"{result['abnormal_count']}/{result['total_companies']} = {result['abnormal_count'] / max(result['total_companies'], 1) * 100:.1f}%"],
        ['API(企业列表)', '/credit/tendering/corps?industryId=2'],
        ['API(资质信息)', '/credit/tendering/qualify/{企业名称}'],
        ['API(异常详情)', '/credit/tendering-extract/qualify'],
        ['判断方式', 'qualifyList 中 abnormal=true'],
        ['生成时间', result['scrape_time']],
    ]
    for i, (k, v) in enumerate(stats_data):
        cell = ws3.cell(row=i + 1, column=1, value=k)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws3.cell(row=i + 1, column=2, value=str(v)).border = thin_border
        ws3.cell(row=i + 1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    
    wb.save(output_file)
    print(f'  [OK] Excel 已保存: {output_file}')
    
    # 同时保存 JSON
    json_file = os.path.join(output_dir, f'{prefix}_{timestamp}.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'  [OK] JSON 已保存: {json_file}')
    
    # 保存最新版本（固定文件名，方便程序读取）
    latest_excel = os.path.join(output_dir, f'{prefix}_最新.xlsx')
    latest_json = os.path.join(output_dir, f'{prefix}_最新.json')
    wb.save(latest_excel)
    with open(latest_json, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'  [OK] 最新版本已更新: {latest_excel}')
    
    return output_file, latest_excel


# ============================================================
# 变化检测
# ============================================================

def load_historical_baseline(output_dir, prefix, days_ago=7):
    """
    加载 N 天前的历史数据作为对比基准。
    扫描 output 目录中时间戳 JSON 文件，找到最接近 N 天前的那一份。
    如果没有任何历史文件，返回 None。
    """
    target_time = datetime.now() - timedelta(days=days_ago)

    candidates = []
    for fname in os.listdir(output_dir):
        if not (fname.startswith(prefix) and fname.endswith('.json')):
            continue
        if '_最新' in fname:
            continue
        # 从文件名提取时间戳：质量检测_资质异常_YYYYMMDD_HHMMSS.json
        base = fname[len(prefix) + 1:-5]  # 去掉前缀+下划线 和 .json
        try:
            file_time = datetime.strptime(base, '%Y%m%d_%H%M%S')
        except ValueError:
            continue
        # 只考虑 target_time 前后的文件（允许最大偏移14天）
        diff = abs((file_time - target_time).total_seconds())
        if diff <= 14 * 86400:
            candidates.append((diff, file_time, fname))

    if not candidates:
        # 没有 7 天附近的数据，尝试使用最早的历史文件
        all_files = []
        for fname in os.listdir(output_dir):
            if not (fname.startswith(prefix) and fname.endswith('.json')):
                continue
            if '_最新' in fname:
                continue
            base = fname[len(prefix) + 1:-5]
            try:
                file_time = datetime.strptime(base, '%Y%m%d_%H%M%S')
                all_files.append((file_time, fname))
            except ValueError:
                continue
        if all_files:
            all_files.sort()
            oldest_time, oldest_fname = all_files[0]
            print(f'\n[变化检测] 无 7 天附近的历史数据，使用最早的历史文件: {oldest_fname}')
            filepath = os.path.join(output_dir, oldest_fname)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return None
        print('\n[变化检测] 无历史数据，首次运行')
        return None

    # 选取最接近 target_time 的文件
    candidates.sort()
    _, best_time, best_fname = candidates[0]
    days_diff = (target_time - best_time).days
    direction = "后" if days_diff < 0 else "前"
    print(f'\n[变化检测] 对比基准: {best_fname}（{abs(days_diff)} 天{direction}）')

    filepath = os.path.join(output_dir, best_fname)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f'[变化检测] 读取历史文件失败: {e}')
        return None


def check_whats_changed(old_result, new_result):
    """对比新旧结果，返回变化描述"""
    if not old_result:
        return None, "首次运行，无历史数据对比"
    
    old_companies = {c['companyName']: c for c in old_result.get('abnormal_companies', [])}
    new_companies = {c['companyName']: c for c in new_result.get('abnormal_companies', [])}
    
    old_names = set(old_companies.keys())
    new_names = set(new_companies.keys())
    messages = []

    # 新增异常企业
    added = new_names - old_names
    if added:
        messages.append(f"🆕 新增 {len(added)} 家异常企业：")
        for name in added:
            items = new_companies[name]['abnormalItems']
            reasons = '；'.join(item.get('qualifyName', '') for item in items)
            messages.append(f"   - {name}（{len(items)} 项：{reasons}）")

    # 恢复正常企业
    removed = old_names - new_names
    if removed:
        messages.append(f"✅ {len(removed)} 家企业恢复正常：")
        for name in removed:
            messages.append(f"   - {name}")

    # 仍在异常列表中的企业（检查异常项是否变化）
    common = old_names & new_names
    changed = []
    for name in common:
        old_items = {item.get('qualifyName', '') for item in old_companies[name]['abnormalItems']}
        new_items = {item.get('qualifyName', '') for item in new_companies[name]['abnormalItems']}
        if old_items != new_items:
            changed.append(name)
    
    if changed:
        messages.append(f"⚠️ {len(changed)} 家企业异常项发生变化：")
        for name in changed:
            messages.append(f"   - {name}")

    if not messages:
        return {
            "messages": [],
            "no_change_msg": "[OK] 无变化（异常企业列表与上次相同）",
            "companies": {"added": [], "removed": [], "changed": []}
        }

    return {
        "messages": messages,
        "no_change_msg": "",
        "companies": {
            "added": sorted(list(added)),
            "removed": sorted(list(removed)),
            "changed": sorted(changed)
        }
    }


def save_change_history(config, result):
    """将变化检测结果保存到历史文件，自动清理超过 7 天的记录"""
    messages = result.get("messages", []) if isinstance(result, dict) else (result or [])
    no_change_msg = result.get("no_change_msg", "") if isinstance(result, dict) else ""
    companies = result.get("companies", {}) if isinstance(result, dict) else {"added": [], "removed": [], "changed": []}
    output_dir = config["output"]["output_dir"]
    history_file = os.path.join(output_dir, "change_history.json")

    # 加载已有历史
    history = {"changes": []}
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except Exception:
            pass

    # 构建新条目
    entry = {
        "time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "has_change": bool(messages),
        "messages": messages or [],
        "companies": companies,
        "no_change_msg": no_change_msg or ""
    }

    # 统计变化类型
    if messages:
        stats = {"added": 0, "removed": 0, "changed": 0}
        for msg in messages:
            if msg.startswith('🆕'):
                m_count = re.search(r'新增\s*(\d+)\s*家', msg)
                if m_count:
                    stats["added"] = int(m_count.group(1))
            elif msg.startswith('✅'):
                m_count = re.search(r'(\d+)\s*家', msg)
                if m_count:
                    stats["removed"] = int(m_count.group(1))
            elif msg.startswith('⚠️'):
                m_count = re.search(r'(\d+)\s*家', msg)
                if m_count:
                    stats["changed"] = int(m_count.group(1))
        entry["stats"] = stats

    history["changes"].append(entry)

    # 只保留最近 7 天
    cutoff = datetime.now() - timedelta(days=7)
    history["changes"] = [
        c for c in history["changes"]
        if datetime.strptime(c["time"], '%Y-%m-%d %H:%M:%S') >= cutoff
    ]

    # 保存
    with open(history_file, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

    print(f'\n[变化历史] 已保存到 {history_file}，当前共 {len(history["changes"])} 条记录（7天内）')


# ============================================================
# 定时调度
# ============================================================

def run_schedule(config):
    """内置调度器：每天指定时间运行"""
    try:
        import schedule
    except ImportError:
        print('  ⚠ 未安装 schedule 库，正在安装...')
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'schedule'])
        import schedule
    
    schedule_time = config["schedule"]["time"]
    
    def job():
        print(f'\n{"=" * 70}')
        print(f'  定时任务触发: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        print(f'{"=" * 70}')
        result = run_scrape(config)
        if result:
            # 变化检测：与 7 天前的历史数据对比
            output_dir = config["output"]["output_dir"]
            prefix = config["output"]["filename_prefix"]
            old_result = load_historical_baseline(output_dir, prefix, days_ago=7)

            # 生成 Excel
            output_file, latest_file = generate_excel(result, config)

            # 变化检测
            change_result = check_whats_changed(old_result, result)
            change_messages = change_result["messages"]
            no_change_msg = change_result["no_change_msg"]
            if change_messages:
                print('\n变化检测：')
                for msg in change_messages:
                    print(f'  {msg}')
            elif no_change_msg:
                print(f'\n变化检测：{no_change_msg}')

            # 保存变化历史（保留7天）
            save_change_history(config, change_result)

            # 发送微信通知
            print('\n[通知] 发送微信通知...')
            notify.send_notification(config, result, change_messages, no_change_msg)
        else:
            # 抓取失败也发通知
            print('\n[通知] 抓取失败，发送告警通知...')
            fail_result = {
                'scrape_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_companies': 0,
                'normal_count': 0,
                'abnormal_count': 0,
                'error_count': 1,
                'abnormal_companies': [],
            }
            fail_title = "质量检测资质异常监控 - 抓取失败告警"
            fail_content = f"## 抓取失败告警\n\n定时任务在 {fail_result['scrape_time']} 运行失败。\n\n请检查网络连接和程序日志。"
            notif_config = config.get('notification', {})
            if notif_config.get('enabled', False):
                sc = notif_config.get('serverchan', {})
                wx = notif_config.get('wxpusher', {})
                if sc.get('enabled'):
                    notify.send_serverchan(sc.get('sendkey', ''), fail_title, fail_content)
                if wx.get('enabled'):
                    notify.send_wxpusher(
                        wx.get('app_token', ''), fail_title, fail_content,
                        wx.get('topic_ids', []), wx.get('uids', [])
                    )
                
    schedule.every().day.at(schedule_time).do(job)
    print(f'\n[OK] 定时任务已设置：每天 {schedule_time} 运行')
    print(f'  按 Ctrl+C 停止\n')
    
    # 立即运行一次
    print('立即运行一次...')
    job()
    
    # 循环等待
    while True:
        schedule.run_pending()
        time.sleep(60)


# ============================================================
# 主程序
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='福建省建设行业平台 - 质量检测企业资质异常自动统计')
    parser.add_argument('--config', default='config.ini', help='配置文件路径')
    parser.add_argument('--schedule', action='store_true', help='启用定时运行模式')
    parser.add_argument('--time', help='定时运行时间 (格式: HH:MM，默认 08:00)')
    parser.add_argument('--output-dir', help='输出目录（覆盖配置文件）')
    args = parser.parse_args()
    
    # 加载配置
    config = load_config(args.config)
    
    # 命令行参数覆盖配置
    if args.time:
        config["schedule"]["time"] = args.time
    if args.output_dir:
        config["output"]["output_dir"] = args.output_dir
    if args.schedule:
        config["schedule"]["enabled"] = True
    
    # 设置日志
    logger = setup_logging(config["output"]["output_dir"])
    
    print(f'\n配置文件: {args.config}')
    print(f'输出目录: {config["output"]["output_dir"]}')
    print(f'定时运行: {"是 (" + config["schedule"]["time"] + ")" if config["schedule"]["enabled"] else "否"}')
    
    if args.schedule:
        # 定时运行模式（仅当显式传入 --schedule 时启用）
        run_schedule(config)
    else:
        # 立即运行一次
        result = run_scrape(config)
        if result:
            # 变化检测：与 7 天前的历史数据对比
            output_dir = config["output"]["output_dir"]
            prefix = config["output"]["filename_prefix"]
            old_result = load_historical_baseline(output_dir, prefix, days_ago=7)

            # 生成 Excel
            output_file, latest_file = generate_excel(result, config)

            # 变化检测
            change_result = check_whats_changed(old_result, result)
            change_messages = change_result["messages"]
            no_change_msg = change_result["no_change_msg"]
            if change_messages:
                print('\n变化检测：')
                for msg in change_messages:
                    print(f'  {msg}')
            elif no_change_msg:
                print(f'\n变化检测：{no_change_msg}')

            # 保存变化历史（保留7天）
            save_change_history(config, change_result)

            print(f'\n{"=" * 70}')
            print(f'  抓取完成')
            print(f'{"=" * 70}')
            print(f'  质量检测企业总数: {result["total_companies"]}')
            print(f'  资质异常企业数: {result["abnormal_count"]}')
            print(f'  资质异常项数: {sum(len(c["abnormalItems"]) for c in result["abnormal_companies"])}')

            if result["abnormal_companies"]:
                print(f'\n  异常企业列表:')
                for comp in result["abnormal_companies"]:
                    items = comp['abnormalItems']
                    print(f'    - {comp["companyName"]}: {len(items)} 项异常')
                    for item in items:
                        print(f'        - {item["qualifyName"]}')

            print(f'\n  报告已保存:')
            print(f'    Excel: {output_file}')
            print(f'    最新: {latest_file}')

            # 发送微信通知
            print(f'\n[通知] 发送微信通知...')
            notify_results = notify.send_notification(config, result, change_messages, no_change_msg)
            if notify_results:
                for channel, success, msg in notify_results:
                    status = '成功' if success else '失败'
                    print(f'    {channel}: {status} - {msg}')
        else:
            print('抓取失败！')


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('\n\n用户中断，程序退出。')
        sys.exit(0)
    except Exception as e:
        print(f'\n\n程序异常: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)
