# -*- coding: utf-8 -*-
"""
资质人员获取 Blueprint
挂载路径: /api/tech_staff

功能：从福建省工程质量检测信用系统搜索企业，
      获取技术人员数据，生成分类汇总 Excel 并支持下载
"""
import os
import io
import time
from datetime import datetime

import requests
import urllib3
from flask import Blueprint, request, jsonify, send_file

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

bp = Blueprint('tech_staff', __name__, url_prefix='/api/tech_staff')


def get_info():
    """模块信息（供 /modules 接口使用）"""
    return {
        'name': '技术人员',
        'desc': '搜索企业，查看技术人员清单，导出带分类统计的Excel报告',
        'icon': '👥',
        'page': '/pages/tech_staff/tech_staff',
        'url_prefix': '/api/tech_staff',
        'bg': '#E65100'
    }


# ============================================================
# 配置
# ============================================================
BASE_URL = "https://220.160.52.164:8813"

# ============================================================
# 样式常量
# ============================================================
BLUE = "4472C4"
DARK_BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
ORANGE = "ED7D31"
GREEN = "70AD47"
YELLOW = "FFC000"

TITLE_ORDER = ["正高级工程师", "高级工程师", "工程师", "助理工程师", "技术员", "未知"]
EDU_ORDER = ["博士研究生", "硕士研究生", "本科", "大专", "中专", "高中", "未知"]


# ============================================================
# 网络会话
# ============================================================
def create_session():
    s = requests.Session()
    s.verify = False
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": BASE_URL,
        "Referer": f"{BASE_URL}/gaia/creditFiles.html",
    })
    return s


# ============================================================
# 核心 API 调用
# ============================================================
def search_companies_api(session, keyword):
    """搜索企业"""
    from urllib.parse import quote
    url = f"{BASE_URL}/credit/publicity-home/publicity?time={int(time.time()*1000)}"
    encoded = quote(keyword) if keyword else ""
    data = f"companyName={encoded}&tradeTypeId=8&pageNum=1&pageSize=50&ministry="

    resp = session.post(url, data=data, timeout=20)
    if resp.status_code == 200:
        result = resp.json()
        code = result.get("code")
        if str(code) in ("1", "200", "0", "success"):
            return result.get("data", {}).get("list", []), result.get("data", {}).get("total", 0)
    return [], 0


def get_technicians_api(session, task_id):
    """分页获取单个提交记录的技术人员"""
    all_techs = []
    page_num = 1
    page_size = 40

    while page_num <= 100:
        url = (f"{BASE_URL}/credit/publicity-detect/checks/{task_id}"
               f"?pageNum={page_num}&pageSize={page_size}&taskId={task_id}"
               f"&time={int(time.time()*1000)}")
        try:
            resp = session.get(url, timeout=20)
            if resp.status_code != 200:
                break
            result = resp.json()
            data_obj = result.get("data", {})
            tech_list = data_obj.get("list", [])
            api_total = data_obj.get("total", 0)
            api_pages = data_obj.get("pages", 0)

            if not tech_list:
                break
            all_techs.extend(tech_list)

            if api_pages > 0 and page_num >= api_pages:
                break
            if api_total > 0 and len(all_techs) >= api_total:
                break
            page_num += 1
        except Exception:
            break
    return all_techs


def get_all_technicians(session, companies):
    """获取所有历史提交记录的技术人员（合并去重）"""
    all_techs = []
    for c in companies:
        task_id = c.get("taskId") or c.get("id")
        if not task_id:
            continue
        techs = get_technicians_api(session, task_id)
        all_techs.extend(techs)
    return all_techs


def _parse_tech(t):
    """统一提取字段"""
    return {
        "name": t.get("name") or t.get("personName") or "",
        "id_card": t.get("identificationNo") or t.get("idNumber") or "",
        "title": t.get("titleLevelName") or t.get("title") or "",
        "cert_no": t.get("titleCertificateNo") or t.get("certificateNo") or "",
        "major": t.get("qualifySpecialtyName") or t.get("major") or "",
        "edu": t.get("educationName") or t.get("education") or "",
    }


# ============================================================
# Excel 导出
# ============================================================
def _sort_key_title(name):
    try:
        return TITLE_ORDER.index(name)
    except ValueError:
        return 99


def _sort_key_edu(name):
    try:
        return EDU_ORDER.index(name)
    except ValueError:
        return 99


def export_excel(company_name, techs):
    """生成 6 个工作表的 Excel 文件，返回文件路径"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=BLUE)
    title_fill = PatternFill("solid", fgColor=DARK_BLUE)
    sub_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    zebra_fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    header_font = Font(bold=True, color=WHITE, size=11)
    title_font = Font(bold=True, size=14, color=WHITE)
    group_font = Font(bold=True, size=12, color=WHITE)
    subtitle_font = Font(bold=True, size=12, color=DARK_BLUE)
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)

    thin = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    total = len(techs)
    parsed = [_parse_tech(t) for t in techs]

    wb = Workbook()
    wb.remove(wb.active)

    HEADERS = ["序号", "姓名", "证件号码", "职称", "证书编号", "专业", "学历"]
    COL_WIDTHS = {"A": 8, "B": 12, "C": 24, "D": 16, "E": 20, "F": 20, "G": 10}

    # ---- Sheet 1: 技术人员清单 ----
    ws1 = wb.create_sheet("技术人员清单")
    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"{company_name} - 技术人员清单"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"共 {total} 人  |  导出时间: {now_str}"
    ws1["A2"].font = Font(size=10, color="666666")
    ws1["A2"].alignment = center

    for col, h in enumerate(HEADERS, 1):
        c = ws1.cell(3, col, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    ws1.row_dimensions[3].height = 25

    for i, t in enumerate(parsed, 1):
        row = i + 3
        vals = [i, t["name"], t["id_card"], t["title"], t["cert_no"], t["major"], t["edu"]]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row, col, v)
            c.alignment = center if col != 6 else left
            c.border = thin_border
            c.font = normal_font
            if i % 2 == 0:
                c.fill = zebra_fill

    for col, w in COL_WIDTHS.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A4"

    # ---- 通用分组函数 ----
    def build_grouped_sheet(sheet_name, field_key, sort_func, colors):
        ws = wb.create_sheet(sheet_name)
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{company_name} - {sheet_name}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:G2")
        ws["A2"] = f"共 {total} 人  |  导出时间: {now_str}"
        ws["A2"].font = Font(size=10, color="666666")
        ws["A2"].alignment = center

        groups = {}
        for t in parsed:
            k = t[field_key] or "未知"
            groups.setdefault(k, []).append(t)
        sorted_keys = sorted(groups.keys(), key=sort_func)

        current_row = 4
        global_seq = 0
        for idx, gk in enumerate(sorted_keys):
            members = groups[gk]
            members.sort(key=lambda x: (_sort_key_title(x["title"]), x["name"]))

            color = colors[idx % len(colors)]
            ws.merge_cells(f"A{current_row}:G{current_row}")
            c = ws.cell(current_row, 1, f"  {gk}  ({len(members)}人)")
            c.font = group_font
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = left
            ws.row_dimensions[current_row].height = 26
            current_row += 1

            for col, h in enumerate(HEADERS, 1):
                c = ws.cell(current_row, col, h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = thin_border
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            for i, t in enumerate(members, 1):
                global_seq += 1
                vals = [global_seq, t["name"], t["id_card"], t["title"], t["cert_no"], t["major"], t["edu"]]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(current_row, col, v)
                    c.alignment = center if col != 6 else left
                    c.border = thin_border
                    c.font = normal_font
                    if i % 2 == 0:
                        c.fill = zebra_fill
                current_row += 1
            current_row += 1

        for col, w in COL_WIDTHS.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A4"

    # Sheet 2-4: 分类
    build_grouped_sheet("按职称分类", "title", _sort_key_title, [DARK_BLUE, BLUE, GREEN, ORANGE, YELLOW])
    build_grouped_sheet("按专业分类", "major",
                        lambda x: -(sum(1 for t in parsed if (t["major"] or "未知") == x)),
                        [GREEN, ORANGE, BLUE, YELLOW])
    build_grouped_sheet("按学历分类", "edu", _sort_key_edu, [DARK_BLUE, BLUE, GREEN, ORANGE, YELLOW])

    # ---- Sheet 5: 交叉统计 ----
    ws5 = wb.create_sheet("交叉统计")
    ws5.merge_cells("A1:H1")
    ws5["A1"] = f"{company_name} - 交叉统计分析"
    ws5["A1"].font = title_font
    ws5["A1"].fill = title_fill
    ws5["A1"].alignment = center
    ws5.row_dimensions[1].height = 32

    titles_list = sorted(set(t["title"] or "未知" for t in parsed), key=_sort_key_title)
    majors_list = sorted(set(t["major"] or "未知" for t in parsed),
                         key=lambda x: -(sum(1 for t in parsed if (t["major"] or "未知") == x)))
    edus_list = sorted(set(t["edu"] or "未知" for t in parsed), key=_sort_key_edu)

    def write_cross_table(ws, start_row, table_title, row_labels, col_labels, row_field, col_field):
        ws.cell(start_row, 1, table_title).font = subtitle_font
        ws.row_dimensions[start_row].height = 24
        hr = start_row + 1

        corner = ws.cell(hr, 1, f"{row_field} \\ {col_field}")
        corner.font = header_font
        corner.fill = header_fill
        corner.alignment = center
        corner.border = thin_border
        for j, cl in enumerate(col_labels, 2):
            c = ws.cell(hr, j, cl)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin_border
        tc = len(col_labels) + 2
        c = ws.cell(hr, tc, "合计")
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = center
        c.border = thin_border
        ws.row_dimensions[hr].height = 22

        for i, rl in enumerate(row_labels, 1):
            r = hr + i
            c = ws.cell(r, 1, rl)
            c.font = bold_font
            c.fill = sub_fill
            c.alignment = left
            c.border = thin_border
            row_total = 0
            for j, cl in enumerate(col_labels, 2):
                cnt = sum(1 for t in parsed if (t[row_field] or "未知") == rl and (t[col_field] or "未知") == cl)
                c = ws.cell(r, j, cnt if cnt > 0 else "")
                c.alignment = center
                c.border = thin_border
                c.font = normal_font
                if cnt > 0:
                    row_total += cnt
                    if cnt >= 10:
                        c.fill = PatternFill("solid", fgColor="FF9999")
                    elif cnt >= 5:
                        c.fill = PatternFill("solid", fgColor="FFCC99")
            c = ws.cell(r, tc, row_total)
            c.font = bold_font
            c.alignment = center
            c.border = thin_border
            c.fill = sub_fill

        sr = hr + len(row_labels) + 1
        c = ws.cell(sr, 1, "合计")
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = center
        c.border = thin_border
        gt = 0
        for j, cl in enumerate(col_labels, 2):
            ct = sum(1 for t in parsed if (t[col_field] or "未知") == cl)
            c = ws.cell(sr, j, ct)
            c.font = bold_font
            c.fill = sub_fill
            c.alignment = center
            c.border = thin_border
            gt += ct
        c = ws.cell(sr, tc, gt)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = center
        c.border = thin_border
        return sr

    end1 = write_cross_table(ws5, 3, "一、职称 x 专业", titles_list, majors_list, "title", "major")
    write_cross_table(ws5, end1 + 3, "二、职称 x 学历", titles_list, edus_list, "title", "edu")

    ws5.column_dimensions["A"].width = 18
    for j in range(2, 10):
        ws5.column_dimensions[get_column_letter(j)].width = 16

    # ---- Sheet 6: 图表汇总 ----
    ws6 = wb.create_sheet("图表汇总")
    ws6.merge_cells("A1:H1")
    ws6["A1"] = f"{company_name} - 图表汇总"
    ws6["A1"].font = title_font
    ws6["A1"].fill = title_fill
    ws6["A1"].alignment = center
    ws6.row_dimensions[1].height = 32

    title_stats, major_stats, edu_stats = {}, {}, {}
    for t in parsed:
        ti = t["title"] or "未知"
        ma = t["major"] or "未知"
        ed = t["edu"] or "未知"
        title_stats[ti] = title_stats.get(ti, 0) + 1
        major_stats[ma] = major_stats.get(ma, 0) + 1
        edu_stats[ed] = edu_stats.get(ed, 0) + 1

    def write_chart_data(ws, start_row, label, data):
        ws.cell(start_row, 1, label).font = header_font
        ws.cell(start_row, 1).fill = header_fill
        ws.cell(start_row, 1).alignment = center
        ws.cell(start_row, 2, "人数").font = header_font
        ws.cell(start_row, 2).fill = header_fill
        ws.cell(start_row, 2).alignment = center
        r = start_row + 1
        for k, v in sorted(data.items(), key=lambda x: -x[1]):
            ws.cell(r, 1, k)
            ws.cell(r, 2, v)
            r += 1
        return r

    r1 = write_chart_data(ws6, 3, "职称", title_stats)
    pie1 = PieChart()
    pie1.title = f"职称分布 (共{total}人)"
    pie1.add_data(Reference(ws6, min_col=2, min_row=3, max_row=r1 - 1), titles_from_data=True)
    pie1.set_categories(Reference(ws6, min_col=1, min_row=4, max_row=r1 - 1))
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent = True
    pie1.dataLabels.showVal = True
    pie1.height = 9
    pie1.width = 14
    ws6.add_chart(pie1, "D3")

    bs = r1 + 3
    r2 = write_chart_data(ws6, bs, "专业", major_stats)
    bar1 = BarChart()
    bar1.type = "bar"
    bar1.title = "专业分布"
    bar1.add_data(Reference(ws6, min_col=2, min_row=bs, max_row=r2 - 1), titles_from_data=True)
    bar1.set_categories(Reference(ws6, min_col=1, min_row=bs + 1, max_row=r2 - 1))
    bar1.dataLabels = DataLabelList()
    bar1.dataLabels.showVal = True
    bar1.height = 9
    bar1.width = 14
    ws6.add_chart(bar1, "D" + str(bs))

    es = r2 + 3
    r3 = write_chart_data(ws6, es, "学历", edu_stats)
    pie2 = PieChart()
    pie2.title = "学历分布"
    pie2.add_data(Reference(ws6, min_col=2, min_row=es, max_row=r3 - 1), titles_from_data=True)
    pie2.set_categories(Reference(ws6, min_col=1, min_row=es + 1, max_row=r3 - 1))
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent = True
    pie2.dataLabels.showVal = True
    pie2.height = 9
    pie2.width = 14
    ws6.add_chart(pie2, "D" + str(es))

    for col, w in {"A": 18, "B": 10}.items():
        ws6.column_dimensions[col].width = w

    # 保存到 output 目录
    bp_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.normpath(os.path.join(bp_dir, '..', 'output'))
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"技术人员_{company_name}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath, filename


# ============================================================
# API 端点
# ============================================================

@bp.route('/search', methods=['GET'])
def search():
    """搜索企业（返回公司列表）"""
    keyword = request.args.get('keyword', '').strip()
    if not keyword:
        return jsonify({'code': -1, 'msg': '请输入搜索关键词'})

    session = create_session()
    try:
        session.get(f"{BASE_URL}/gaia/creditFiles.html", timeout=10)
        companies, total = search_companies_api(session, keyword)
    except Exception as e:
        return jsonify({'code': -1, 'msg': f'连接失败：{str(e)}', 'data': []})

    if not companies:
        return jsonify({'code': 0, 'msg': f'未找到包含"{keyword}"的企业', 'data': [], 'total': 0})

    # 提取关键字段
    result = []
    for c in companies:
        result.append({
            'company_name': c.get('companyName') or c.get('name') or '未知',
            'task_id': c.get('taskId') or c.get('id') or '',
            'receive_time': c.get('receiveTime') or '',
            'code': c.get('code', ''),
            'node_type': c.get('nodeType', ''),
        })

    return jsonify({'code': 0, 'data': result, 'total': total})


@bp.route('/technicians', methods=['GET'])
def get_technicians():
    """获取技术人员数据（含统计汇总）"""
    task_id = request.args.get('taskId', '').strip()
    company_name = request.args.get('companyName', '').strip()

    if not task_id:
        return jsonify({'code': -1, 'msg': '缺少 taskId 参数'})
    if not company_name:
        company_name = '企业'

    session = create_session()
    try:
        session.get(f"{BASE_URL}/gaia/creditFiles.html", timeout=10)
        techs_raw = get_technicians_api(session, task_id)
    except Exception as e:
        return jsonify({'code': -1, 'msg': f'获取失败：{str(e)}'})

    if not techs_raw:
        return jsonify({'code': 0, 'msg': '该公司暂无技术人员记录', 'data': None})

    parsed = [_parse_tech(t) for t in techs_raw]

    # 统计
    title_stats, major_stats, edu_stats = {}, {}, {}
    for t in parsed:
        ti = t['title'] or '未知'
        ma = t['major'] or '未知'
        ed = t['edu'] or '未知'
        title_stats[ti] = title_stats.get(ti, 0) + 1
        major_stats[ma] = major_stats.get(ma, 0) + 1
        edu_stats[ed] = edu_stats.get(ed, 0) + 1

    # 排序统计
    def sort_by_count(stats):
        return sorted(stats.items(), key=lambda x: -x[1])

    return jsonify({
        'code': 0,
        'data': {
            'company_name': company_name,
            'task_id': task_id,
            'total': len(parsed),
            'technicians': parsed,
            # 统计
            'stats': {
                'title': [{'name': k, 'count': v} for k, v in sort_by_count(title_stats)],
                'major': [{'name': k, 'count': v} for k, v in sort_by_count(major_stats)],
                'education': [{'name': k, 'count': v} for k, v in sort_by_count(edu_stats)],
            }
        }
    })


@bp.route('/export', methods=['GET'])
def export():
    """生成并下载 Excel 文件"""
    task_id = request.args.get('taskId', '').strip()
    company_name = request.args.get('companyName', '').strip()

    if not task_id:
        return jsonify({'code': -1, 'msg': '缺少 taskId 参数'})
    if not company_name:
        company_name = '企业'

    session = create_session()
    try:
        session.get(f"{BASE_URL}/gaia/creditFiles.html", timeout=10)
        techs_raw = get_technicians_api(session, task_id)
    except Exception as e:
        return jsonify({'code': -1, 'msg': f'获取失败：{str(e)}'})

    if not techs_raw:
        return jsonify({'code': -1, 'msg': '该公司暂无技术人员记录'})

    try:
        filepath, filename = export_excel(company_name, techs_raw)
    except Exception as e:
        return jsonify({'code': -1, 'msg': f'Excel生成失败：{str(e)}'})

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
